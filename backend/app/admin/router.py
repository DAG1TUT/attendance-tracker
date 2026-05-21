from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, time as dt_time, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from typing import Annotated
from zoneinfo import ZoneInfo

import asyncpg
import openpyxl
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse

from app.admin.schemas import (
    AdminLogOut,
    AttendanceDayEdit,
    DayCell,
    DaySchedule,
    EmployeeCreate,
    EmployeeOut,
    EmployeeUpdate,
    PermissionsUpdate,
    RevenueOut,
    RevenueUpsert,
    SalaryEntry,
    SalaryReport,
    ScheduleEntry,
    ScheduleSession,
    TodayEntry,
    TodayStats,
    WeekSchedule,
    WeekScheduleEntry,
)
from app.auth.service import hash_password
from app.config import settings
from app.dependencies import AdminOnly, OwnerOnly, get_db

router = APIRouter(prefix="/admin", tags=["admin"])


# ── Pending approvals ─────────────────────────────────────────────────────────

@router.get("/employees/pending", response_model=list[EmployeeOut])
async def list_pending(
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    _: Annotated[dict, AdminOnly],
) -> list[dict]:
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """SELECT id, phone, name, role, position, is_active, status,
                      hourly_rate, bonus_percent, created_at, is_owner, permissions
               FROM users WHERE status = 'pending' ORDER BY created_at"""
        )
    return [dict(r) for r in rows]


@router.post("/employees/{employee_id}/approve", response_model=EmployeeOut)
async def approve_employee(
    employee_id: int,
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    _: Annotated[dict, AdminOnly],
) -> dict:
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            """UPDATE users SET status = 'active'
               WHERE id = $1 AND status = 'pending'
               RETURNING id, phone, name, role, position, is_active, status,
                         hourly_rate, bonus_percent, created_at, is_owner, permissions""",
            employee_id,
        )
    if not row:
        raise HTTPException(status_code=404, detail="Сотрудник не найден или уже подтверждён")
    return dict(row)


@router.post("/employees/{employee_id}/reject")
async def reject_employee(
    employee_id: int,
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    _: Annotated[dict, AdminOnly],
) -> dict:
    async with db.acquire() as conn:
        result = await conn.execute(
            "DELETE FROM users WHERE id = $1 AND status = 'pending'", employee_id
        )
    if result == "DELETE 0":
        raise HTTPException(status_code=404, detail="Сотрудник не найден")
    return {"message": "Заявка отклонена"}


# ── Employees ─────────────────────────────────────────────────────────────────

@router.get("/employees", response_model=list[EmployeeOut])
async def list_employees(
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    _: Annotated[dict, AdminOnly],
    is_active: bool | None = None,
) -> list[dict]:
    conditions = ["status = 'active'"]
    args: list = []
    if is_active is not None:
        conditions.append(f"is_active = ${len(args) + 1}")
        args.append(is_active)
    where = "WHERE " + " AND ".join(conditions)
    query = f"""SELECT id, phone, name, role, position, is_active, status,
                       hourly_rate, bonus_percent, created_at, is_owner, permissions
                FROM users {where} ORDER BY name"""
    async with db.acquire() as conn:
        rows = await conn.fetch(query, *args)
    return [dict(r) for r in rows]


@router.post("/employees", status_code=status.HTTP_201_CREATED, response_model=EmployeeOut)
async def create_employee(
    body: EmployeeCreate,
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    _: Annotated[dict, AdminOnly],
) -> dict:
    async with db.acquire() as conn:
        existing = await conn.fetchval("SELECT id FROM users WHERE phone = $1", body.phone)
        if existing:
            raise HTTPException(status_code=409, detail="Телефон уже занят")

        row = await conn.fetchrow(
            """INSERT INTO users (phone, name, password_hash, role, position, hourly_rate, bonus_percent, status)
               VALUES ($1, $2, $3, $4, $5, $6, $7, 'active')
               RETURNING id, phone, name, role, position, is_active, status,
                         hourly_rate, bonus_percent, created_at, is_owner, permissions""",
            body.phone, body.name, hash_password(body.password),
            body.role, body.position, body.hourly_rate, body.bonus_percent,
        )
    return dict(row)


@router.patch("/employees/{employee_id}", response_model=EmployeeOut)
async def update_employee(
    employee_id: int,
    body: EmployeeUpdate,
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    _: Annotated[dict, AdminOnly],
) -> dict:
    async with db.acquire() as conn:
        user = await conn.fetchrow("SELECT id FROM users WHERE id = $1", employee_id)
        if not user:
            raise HTTPException(status_code=404, detail="Сотрудник не найден")

        updates, args = [], []
        idx = 1
        for field, value in {
            "name": body.name, "phone": body.phone, "role": body.role,
            "position": body.position, "is_active": body.is_active,
            "hourly_rate": body.hourly_rate, "bonus_percent": body.bonus_percent,
        }.items():
            if value is not None:
                updates.append(f"{field} = ${idx}"); args.append(value); idx += 1
        if body.password is not None:
            updates.append(f"password_hash = ${idx}"); args.append(hash_password(body.password)); idx += 1

        if not updates:
            raise HTTPException(status_code=400, detail="Нет данных для обновления")

        args.append(employee_id)
        row = await conn.fetchrow(
            f"""UPDATE users SET {', '.join(updates)} WHERE id = ${idx}
                RETURNING id, phone, name, role, position, is_active, status,
                          hourly_rate, bonus_percent, created_at, is_owner, permissions""",
            *args,
        )
    return dict(row)


@router.delete("/employees/{employee_id}")
async def delete_employee(
    employee_id: int,
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    _: Annotated[dict, AdminOnly],
) -> dict:
    async with db.acquire() as conn:
        result = await conn.execute("DELETE FROM users WHERE id = $1", employee_id)
    if result == "DELETE 0":
        raise HTTPException(status_code=404, detail="Сотрудник не найден")
    return {"message": "Удалено"}


# ── Permissions (owner only) ───────────────────────────────────────────────────

@router.get("/employees/{employee_id}/permissions")
async def get_permissions(
    employee_id: int,
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    _: Annotated[dict, OwnerOnly],
) -> dict:
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id, name, is_owner, permissions FROM users WHERE id = $1", employee_id
        )
    if not row:
        raise HTTPException(404, "Не найден")
    return {
        "user_id": row["id"],
        "name": row["name"],
        "is_owner": row["is_owner"],
        "permissions": row["permissions"],
    }


@router.put("/employees/{employee_id}/permissions")
async def set_permissions(
    employee_id: int,
    body: PermissionsUpdate,
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    current_user: Annotated[dict, OwnerOnly],
) -> dict:
    import json
    async with db.acquire() as conn:
        target = await conn.fetchrow("SELECT id, is_owner FROM users WHERE id = $1", employee_id)
        if not target:
            raise HTTPException(404, "Не найден")
        if target["is_owner"] and target["id"] != current_user["id"]:
            raise HTTPException(403, "Нельзя изменять права другого владельца")

        perms_json = json.dumps(body.permissions) if body.permissions is not None else None
        await conn.execute(
            "UPDATE users SET permissions = $1::jsonb WHERE id = $2",
            perms_json, employee_id,
        )
    return {"message": "Права обновлены"}


@router.post("/employees/{employee_id}/make-owner")
async def make_owner(
    employee_id: int,
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    _: Annotated[dict, OwnerOnly],
) -> dict:
    async with db.acquire() as conn:
        row = await conn.fetchrow("SELECT role FROM users WHERE id = $1", employee_id)
        if not row or row["role"] != "admin":
            raise HTTPException(400, "Только администратор может стать владельцем")
        await conn.execute("UPDATE users SET is_owner = TRUE WHERE id = $1", employee_id)
    return {"message": "Пользователь назначен владельцем"}


# ── Logs ──────────────────────────────────────────────────────────────────────

@router.get("/logs", response_model=list[AdminLogOut])
async def list_logs(
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    _: Annotated[dict, AdminOnly],
    user_id: int | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    suspicious_only: bool = False,
    limit: int = 100,
    offset: int = 0,
) -> list[dict]:
    conditions, args = [], []
    idx = 1
    if user_id is not None:
        conditions.append(f"al.user_id = ${idx}"); args.append(user_id); idx += 1
    if suspicious_only:
        conditions.append("al.is_suspicious = TRUE")
    if date_from:
        conditions.append(f"al.timestamp >= ${idx}::timestamptz"); args.append(date_from); idx += 1
    if date_to:
        conditions.append(f"al.timestamp <= ${idx}::timestamptz"); args.append(date_to); idx += 1

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    args.extend([limit, offset])
    query = f"""
        SELECT al.id, al.user_id, u.name AS employee_name,
               al.action, al.timestamp, al.ip_address,
               al.device_id, al.user_agent, al.is_suspicious, al.suspicious_reason
        FROM attendance_logs al JOIN users u ON u.id = al.user_id
        {where} ORDER BY al.timestamp DESC LIMIT ${idx} OFFSET ${idx + 1}
    """
    async with db.acquire() as conn:
        rows = await conn.fetch(query, *args)
    return [dict(r) for r in rows]


@router.get("/logs/suspicious", response_model=list[AdminLogOut])
async def list_suspicious_logs(
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    _: Annotated[dict, AdminOnly],
    limit: int = 50,
) -> list[dict]:
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """SELECT al.id, al.user_id, u.name AS employee_name,
                      al.action, al.timestamp, al.ip_address,
                      al.device_id, al.user_agent, al.is_suspicious, al.suspicious_reason
               FROM attendance_logs al JOIN users u ON u.id = al.user_id
               WHERE al.is_suspicious = TRUE ORDER BY al.timestamp DESC LIMIT $1""",
            limit,
        )
    return [dict(r) for r in rows]


# ── Today stats ───────────────────────────────────────────────────────────────

@router.get("/stats/today", response_model=TodayStats)
async def stats_today(
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    _: Annotated[dict, AdminOnly],
) -> dict:
    tz = ZoneInfo(settings.timezone)
    shift_total_minutes = settings.shift_start_hour * 60 + settings.late_threshold_minutes

    # Today's date in Moscow timezone
    today_moscow = datetime.now(tz).date()

    async with db.acquire() as conn:
        all_employees = await conn.fetch(
            """SELECT id, name FROM users
               WHERE is_active = TRUE AND role = 'employee' AND status = 'active'
               ORDER BY name"""
        )
        # Find first check-in today (in Moscow time)
        today_checkins = await conn.fetch(
            """SELECT DISTINCT ON (user_id) user_id, timestamp
               FROM attendance_logs
               WHERE action = 'check_in'
                 AND (timestamp AT TIME ZONE $1)::date = $2
               ORDER BY user_id, timestamp ASC""",
            settings.timezone, today_moscow,
        )

    checkin_map = {r["user_id"]: r["timestamp"] for r in today_checkins}
    present, absent, late = [], [], []
    for emp in all_employees:
        uid, name = emp["id"], emp["name"]
        if uid not in checkin_map:
            absent.append(TodayEntry(user_id=uid, name=name))
        else:
            ts: datetime = checkin_map[uid]
            # Convert UTC timestamp to Moscow time for display and late check
            ts_moscow = ts.astimezone(tz)
            checked_in_str = ts_moscow.strftime("%H:%M")
            entry_minutes = ts_moscow.hour * 60 + ts_moscow.minute
            if entry_minutes > shift_total_minutes:
                late.append(TodayEntry(user_id=uid, name=name,
                                       checked_in_at=checked_in_str,
                                       late_minutes=entry_minutes - shift_total_minutes))
            else:
                present.append(TodayEntry(user_id=uid, name=name, checked_in_at=checked_in_str))

    return {"present": present, "absent": absent, "late": late}


# ── Revenue ───────────────────────────────────────────────────────────────────

@router.post("/revenue", response_model=RevenueOut)
async def upsert_revenue(
    body: RevenueUpsert,
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    current_user: Annotated[dict, AdminOnly],
) -> dict:
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            """INSERT INTO revenue_entries (date, amount, note, created_by)
               VALUES ($1, $2, $3, $4)
               ON CONFLICT (date) DO UPDATE
               SET amount = EXCLUDED.amount, note = EXCLUDED.note
               RETURNING id, date, amount, note, created_at""",
            body.date, body.amount, body.note, current_user["id"],
        )
    return dict(row)


@router.get("/revenue", response_model=list[RevenueOut])
async def list_revenue(
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    _: Annotated[dict, AdminOnly],
    date_from: date | None = None,
    date_to: date | None = None,
) -> list[dict]:
    conditions, args = [], []
    idx = 1
    if date_from:
        conditions.append(f"date >= ${idx}"); args.append(date_from); idx += 1
    if date_to:
        conditions.append(f"date <= ${idx}"); args.append(date_to); idx += 1
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    async with db.acquire() as conn:
        rows = await conn.fetch(
            f"SELECT id, date, amount, note, created_at FROM revenue_entries {where} ORDER BY date DESC",
            *args,
        )
    return [dict(r) for r in rows]


@router.get("/revenue/export")
async def export_revenue(
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    _: Annotated[dict, AdminOnly],
    date_from: date | None = None,
    date_to: date | None = None,
) -> StreamingResponse:
    conditions, args = [], []
    idx = 1
    if date_from:
        conditions.append(f"date >= ${idx}"); args.append(date_from); idx += 1
    if date_to:
        conditions.append(f"date <= ${idx}"); args.append(date_to); idx += 1
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    async with db.acquire() as conn:
        rows = await conn.fetch(
            f"SELECT id, date, amount, note, created_at FROM revenue_entries {where} ORDER BY date ASC",
            *args,
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Выручка"

    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="6C63FF")
    header_alignment = Alignment(horizontal="center")

    headers = ["Дата", "Выручка (₽)", "Примечание"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    total_amount = Decimal("0")
    for r in rows:
        ws.append([
            r["date"].strftime("%d.%m.%Y"),
            float(r["amount"]),
            r["note"] or "",
        ])
        total_amount += Decimal(str(r["amount"]))

    # ИТОГО row
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="ИТОГО").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=float(total_amount)).font = Font(bold=True)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 30
    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"revenue_{date_from or 'all'}_{date_to or 'all'}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Salary calculation ────────────────────────────────────────────────────────

def _calc_hours_from_logs(user_logs: dict[int, list], user_id: int) -> float:
    events = user_logs.get(user_id, [])
    total_sec = 0.0
    last_in = None
    for ev in sorted(events, key=lambda e: e["timestamp"]):
        if ev["action"] == "check_in":
            last_in = ev["timestamp"]
        elif ev["action"] == "check_out" and last_in:
            total_sec += (ev["timestamp"] - last_in).total_seconds()
            last_in = None
    return round(total_sec / 3600, 2)


@router.get("/salary", response_model=SalaryReport)
async def calculate_salary(
    date_from: date,
    date_to: date,
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    _: Annotated[dict, AdminOnly],
) -> dict:
    async with db.acquire() as conn:
        employees = await conn.fetch(
            """SELECT id, name, role, position, hourly_rate, bonus_percent
               FROM users
               WHERE is_active = TRUE AND status = 'active' AND role = 'employee'
               ORDER BY name"""
        )
        logs = await conn.fetch(
            """SELECT user_id, action, timestamp
               FROM attendance_logs
               WHERE timestamp >= $1::date
                 AND timestamp < ($2::date + INTERVAL '1 day')
                 AND action IN ('check_in', 'check_out')
               ORDER BY user_id, timestamp ASC""",
            date_from, date_to,
        )
        rev_row = await conn.fetchrow(
            """SELECT COALESCE(SUM(amount), 0) AS total
               FROM revenue_entries WHERE date >= $1 AND date <= $2""",
            date_from, date_to,
        )

    total_revenue = Decimal(str(rev_row["total"]))

    # Group logs by user
    user_logs: dict[int, list] = {}
    for log in logs:
        user_logs.setdefault(log["user_id"], []).append(log)

    result = []
    for emp in employees:
        hours = _calc_hours_from_logs(user_logs, emp["id"])
        rate = Decimal(str(emp["hourly_rate"]))
        pct = Decimal(str(emp["bonus_percent"]))
        base = (Decimal(str(hours)) * rate).quantize(Decimal("0.01"))
        bonus = (total_revenue * pct / Decimal("100")).quantize(Decimal("0.01"))
        result.append(SalaryEntry(
            user_id=emp["id"], name=emp["name"], role=emp["role"],
            position=emp["position"],
            hourly_rate=rate, bonus_percent=pct,
            hours_worked=hours, base_pay=base, bonus_pay=bonus,
            total_pay=(base + bonus),
        ))

    return SalaryReport(
        date_from=date_from, date_to=date_to,
        total_revenue=total_revenue, employees=result,
    )


@router.get("/salary/export")
async def export_salary(
    date_from: date,
    date_to: date,
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    _: Annotated[dict, AdminOnly],
) -> StreamingResponse:
    async with db.acquire() as conn:
        employees = await conn.fetch(
            """SELECT id, name, role, position, hourly_rate, bonus_percent
               FROM users
               WHERE is_active = TRUE AND status = 'active' AND role = 'employee'
               ORDER BY name"""
        )
        logs = await conn.fetch(
            """SELECT user_id, action, timestamp
               FROM attendance_logs
               WHERE timestamp >= $1::date
                 AND timestamp < ($2::date + INTERVAL '1 day')
                 AND action IN ('check_in', 'check_out')
               ORDER BY user_id, timestamp ASC""",
            date_from, date_to,
        )
        rev_row = await conn.fetchrow(
            """SELECT COALESCE(SUM(amount), 0) AS total
               FROM revenue_entries WHERE date >= $1 AND date <= $2""",
            date_from, date_to,
        )

    total_revenue = Decimal(str(rev_row["total"]))
    user_logs: dict[int, list] = {}
    for log in logs:
        user_logs.setdefault(log["user_id"], []).append(log)

    POSITION_RU = {
        "employee": "Сотрудник", "runner": "Ранер",
        "cook": "Повар", "barman": "Бармен", "admin": "Администратор",
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Зарплата"

    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="6C63FF")
    header_alignment = Alignment(horizontal="center")

    headers = ["Имя", "Должность", "Ставка ₽/ч", "Бонус %", "Часы", "Оклад", "Бонус", "Итого"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    total_base = Decimal("0")
    total_bonus_sum = Decimal("0")
    total_all = Decimal("0")

    for emp in employees:
        hours = _calc_hours_from_logs(user_logs, emp["id"])
        rate = Decimal(str(emp["hourly_rate"]))
        pct = Decimal(str(emp["bonus_percent"]))
        base = (Decimal(str(hours)) * rate).quantize(Decimal("0.01"))
        bonus = (total_revenue * pct / Decimal("100")).quantize(Decimal("0.01"))
        total_pay = base + bonus
        total_base += base
        total_bonus_sum += bonus
        total_all += total_pay

        ws.append([
            emp["name"],
            POSITION_RU.get(emp["position"], emp["position"]),
            float(rate),
            float(pct),
            hours,
            float(base),
            float(bonus),
            float(total_pay),
        ])

    # ИТОГО row
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="ИТОГО").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=float(total_base)).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=float(total_bonus_sum)).font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=float(total_all)).font = Font(bold=True)

    col_widths = [22, 18, 12, 10, 8, 12, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"salary_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Daily schedule / Timeline ─────────────────────────────────────────────────

@router.get("/stats/schedule", response_model=DaySchedule)
async def stats_schedule(
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    _: Annotated[dict, AdminOnly],
    target_date: date | None = None,
) -> dict:
    tz = ZoneInfo(settings.timezone)
    day = target_date or datetime.now(tz).date()

    async with db.acquire() as conn:
        employees = await conn.fetch(
            """SELECT id, name, hourly_rate, bonus_percent
               FROM users
               WHERE is_active = TRUE AND status = 'active' AND role = 'employee'
               ORDER BY name"""
        )
        logs = await conn.fetch(
            """SELECT user_id, action, timestamp
               FROM attendance_logs
               WHERE (timestamp AT TIME ZONE $1)::date = $2
                 AND action IN ('check_in', 'check_out')
               ORDER BY user_id, timestamp ASC""",
            settings.timezone, day,
        )
        rev_row = await conn.fetchrow(
            "SELECT COALESCE(amount, 0) AS total FROM revenue_entries WHERE date = $1",
            day,
        )

    revenue = Decimal(str(rev_row["total"])) if rev_row else Decimal("0")

    user_logs: dict[int, list] = {}
    for log in logs:
        user_logs.setdefault(log["user_id"], []).append(log)

    result = []
    now_utc = datetime.now(timezone.utc)

    for emp in employees:
        events = user_logs.get(emp["id"], [])
        sessions: list[ScheduleSession] = []
        last_in: datetime | None = None
        total_seconds = 0.0
        is_active = False

        for ev in events:
            ts: datetime = ev["timestamp"]
            if ev["action"] == "check_in":
                last_in = ts
            elif ev["action"] == "check_out" and last_in:
                minutes = int((ts - last_in).total_seconds() / 60)
                total_seconds += minutes * 60
                sessions.append(ScheduleSession(check_in=last_in, check_out=ts, minutes=minutes))
                last_in = None

        if last_in:
            minutes = int((now_utc - last_in).total_seconds() / 60)
            total_seconds += minutes * 60
            sessions.append(ScheduleSession(check_in=last_in, check_out=None, minutes=minutes))
            is_active = True

        total_hours = round(total_seconds / 3600, 2)
        rate = Decimal(str(emp["hourly_rate"]))
        pct = Decimal(str(emp["bonus_percent"]))
        base = (Decimal(str(total_hours)) * rate).quantize(Decimal("0.01"))
        bonus = (revenue * pct / Decimal("100")).quantize(Decimal("0.01"))

        result.append(ScheduleEntry(
            user_id=emp["id"], name=emp["name"],
            hourly_rate=rate, bonus_percent=pct,
            sessions=sessions, total_hours=total_hours,
            base_pay=base, bonus_pay=bonus, total_pay=(base + bonus),
            is_active=is_active,
        ))

    return DaySchedule(date=day, revenue=revenue, employees=result)


# ── Weekly schedule table ─────────────────────────────────────────────────────

@router.get("/stats/week", response_model=WeekSchedule)
async def stats_week(
    date_from: date,
    date_to: date,
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    _: Annotated[dict, AdminOnly],
) -> dict:
    tz = ZoneInfo(settings.timezone)
    dates = [date_from + timedelta(days=i) for i in range((date_to - date_from).days + 1)]

    async with db.acquire() as conn:
        employees = await conn.fetch(
            """SELECT id, name FROM users
               WHERE is_active = TRUE AND status = 'active' AND role = 'employee'
               ORDER BY name"""
        )
        logs = await conn.fetch(
            """SELECT user_id, action, timestamp
               FROM attendance_logs
               WHERE (timestamp AT TIME ZONE $1)::date >= $2
                 AND (timestamp AT TIME ZONE $1)::date <= $3
                 AND action IN ('check_in', 'check_out')
               ORDER BY user_id, timestamp ASC""",
            settings.timezone, date_from, date_to,
        )

    # Group logs by user_id → date (Moscow) → list
    user_day: dict[int, dict[date, list]] = defaultdict(lambda: defaultdict(list))
    for log in logs:
        ts_moscow = log["timestamp"].astimezone(tz)
        user_day[log["user_id"]][ts_moscow.date()].append(
            {"action": log["action"], "timestamp": log["timestamp"]}
        )

    result = []
    for emp in employees:
        days = []
        for d in dates:
            day_logs = sorted(user_day[emp["id"]].get(d, []), key=lambda x: x["timestamp"])
            check_in_ts = None
            check_out_ts = None
            total_sec = 0.0
            last_in = None
            for ev in day_logs:
                if ev["action"] == "check_in":
                    if check_in_ts is None:
                        check_in_ts = ev["timestamp"]
                    last_in = ev["timestamp"]
                elif ev["action"] == "check_out" and last_in:
                    check_out_ts = ev["timestamp"]
                    total_sec += (ev["timestamp"] - last_in).total_seconds()
                    last_in = None
            days.append(DayCell(
                date=d,
                check_in=check_in_ts,
                check_out=check_out_ts,
                hours=round(total_sec / 3600, 1),
            ))
        result.append(WeekScheduleEntry(user_id=emp["id"], name=emp["name"], days=days))

    return WeekSchedule(date_from=date_from, date_to=date_to, dates=dates, employees=result)


@router.get("/stats/week/export")
async def export_schedule(
    date_from: date,
    date_to: date,
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    _: Annotated[dict, AdminOnly],
) -> StreamingResponse:
    tz = ZoneInfo(settings.timezone)
    dates = [date_from + timedelta(days=i) for i in range((date_to - date_from).days + 1)]

    async with db.acquire() as conn:
        employees = await conn.fetch(
            """SELECT id, name FROM users
               WHERE is_active = TRUE AND status = 'active' AND role = 'employee'
               ORDER BY name"""
        )
        logs = await conn.fetch(
            """SELECT user_id, action, timestamp
               FROM attendance_logs
               WHERE (timestamp AT TIME ZONE $1)::date >= $2
                 AND (timestamp AT TIME ZONE $1)::date <= $3
                 AND action IN ('check_in', 'check_out')
               ORDER BY user_id, timestamp ASC""",
            settings.timezone, date_from, date_to,
        )

    user_day: dict[int, dict[date, list]] = defaultdict(lambda: defaultdict(list))
    for log in logs:
        ts_moscow = log["timestamp"].astimezone(tz)
        user_day[log["user_id"]][ts_moscow.date()].append(
            {"action": log["action"], "timestamp": log["timestamp"]}
        )

    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "График"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="6C63FF")
    header_alignment = Alignment(horizontal="center")

    # Header row: Сотрудник | date1 | date2 | ... | Итого ч
    header = ["Сотрудник"] + [d.strftime("%d.%m") for d in dates] + ["Итого ч"]
    ws.append(header)
    for col_idx, _ in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for emp in employees:
        row_data = [emp["name"]]
        total_hours = 0.0
        for d in dates:
            day_logs = sorted(user_day[emp["id"]].get(d, []), key=lambda x: x["timestamp"])
            check_in_ts = check_out_ts = last_in = None
            total_sec = 0.0
            for ev in day_logs:
                if ev["action"] == "check_in":
                    if check_in_ts is None:
                        check_in_ts = ev["timestamp"]
                    last_in = ev["timestamp"]
                elif ev["action"] == "check_out" and last_in:
                    check_out_ts = ev["timestamp"]
                    total_sec += (ev["timestamp"] - last_in).total_seconds()
                    last_in = None
            hours = round(total_sec / 3600, 1)
            total_hours += hours
            if check_in_ts or check_out_ts:
                in_str = check_in_ts.astimezone(tz).strftime("%H:%M") if check_in_ts else "—"
                out_str = check_out_ts.astimezone(tz).strftime("%H:%M") if check_out_ts else "—"
                cell_val = f"{in_str} - {out_str} / {hours}ч"
            else:
                cell_val = "—"
            row_data.append(cell_val)
        row_data.append(round(total_hours, 1))
        ws.append(row_data)

    ws.column_dimensions["A"].width = 22
    for i, _ in enumerate(dates, 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20
    ws.column_dimensions[openpyxl.utils.get_column_letter(len(dates) + 2)].width = 10
    ws.freeze_panes = "B2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"schedule_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Edit attendance day (admin manual edit) ───────────────────────────────────

@router.put("/attendance/day")
async def edit_attendance_day(
    body: AttendanceDayEdit,
    db: Annotated[asyncpg.Pool, Depends(get_db)],
    _: Annotated[dict, AdminOnly],
) -> dict:
    tz = ZoneInfo(settings.timezone)

    async with db.acquire() as conn:
        # Verify employee exists
        emp = await conn.fetchrow("SELECT id FROM users WHERE id = $1", body.user_id)
        if not emp:
            raise HTTPException(status_code=404, detail="Сотрудник не найден")

        # Delete existing logs for this user on this date (Moscow)
        await conn.execute(
            """DELETE FROM attendance_logs
               WHERE user_id = $1
                 AND (timestamp AT TIME ZONE $2)::date = $3""",
            body.user_id, settings.timezone, body.date,
        )

        if body.check_in:
            naive = datetime.combine(body.date, dt_time.fromisoformat(body.check_in))
            ts = naive.replace(tzinfo=tz)
            await conn.execute(
                """INSERT INTO attendance_logs
                   (user_id, action, timestamp, ip_address, device_id, user_agent)
                   VALUES ($1, 'check_in', $2, 'manual', 'admin-edit', 'admin-edit')""",
                body.user_id, ts,
            )

        if body.check_out:
            naive = datetime.combine(body.date, dt_time.fromisoformat(body.check_out))
            ts = naive.replace(tzinfo=tz)
            await conn.execute(
                """INSERT INTO attendance_logs
                   (user_id, action, timestamp, ip_address, device_id, user_agent)
                   VALUES ($1, 'check_out', $2, 'manual', 'admin-edit', 'admin-edit')""",
                body.user_id, ts,
            )

    return {"message": "Обновлено"}
